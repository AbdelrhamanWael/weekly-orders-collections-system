# ============================================================
#  Project  : Weekly Orders & Collections Reconciliation System
#             نظام ربط الطلبات والتحصيل الأسبوعي
#  Module   : generate_report.py — Excel Report Generator
#  Developer : Abdelrhaman Wael Mohammed
#  Email     : abdelrhamanwael8@gmail.com
#  LinkedIn  : linkedin.com/in/abdelrhaman-wael-mohammed-790171366
#  Created   : February 2026
# ============================================================
import sqlite3
import pandas as pd
import os
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB_NAME = "finance_system.db"
OUTPUT_DIR = "reports"

def get_db_connection():
    return sqlite3.connect(DB_NAME)

def generate_weekly_report(snapshot_id=0, label=''):
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
        
    conn = get_db_connection()
    
    # Query to join Orders and Collections, including cost fields for Net Profit
    # Added LEFT JOIN for physical_returns to track physical receipt
    query = f'''
    SELECT 
        o.order_id,
        o.platform,
        o.account_name,
        o.order_date,
        o.week_number,
        o.price                                          AS expected_amount,
        o.cost,
        o.shipping,
        o.commission,
        o.tax,
        o.items_summary,
        o.payment_method,
        o.salla_status,
        COALESCE(SUM(CASE WHEN c.is_return=0 THEN c.original_amount  ELSE 0 END), 0) AS original_collected,
        COALESCE(SUM(CASE WHEN c.is_return=0 THEN c.collection_fee   ELSE 0 END), 0) AS total_collection_fee,
        COALESCE(SUM(CASE WHEN c.is_return=0 THEN c.collected_amount ELSE 0 END), 0) AS collected_amount,
        COALESCE(SUM(CASE WHEN c.is_return=1 THEN ABS(c.collected_amount) ELSE 0 END), 0) AS returned_amount,
        MAX(c.collection_date)                           AS last_collection_date,
        COUNT(CASE WHEN c.is_return=0 THEN 1 END)        AS transaction_count,
        COUNT(CASE WHEN c.is_return=1 THEN 1 END)        AS return_count,
        MAX(c.account_name)                              AS collection_account,
        MAX(CASE WHEN pr.id IS NOT NULL THEN 1 ELSE 0 END) AS physical_receipt_flag
    FROM orders o
    LEFT JOIN collections c
        ON o.order_id = c.order_id AND c.snapshot_id = {snapshot_id}
    LEFT JOIN physical_returns pr
        ON o.order_id = pr.tracking_id OR o.tracking_number = pr.tracking_id
    WHERE o.snapshot_id = {snapshot_id}
    GROUP BY o.order_id
    ORDER BY o.platform, o.account_name, o.order_date
    '''
    
    df = pd.read_sql_query(query, conn)
    conn.close()
    
    # حساب صافي الربح: المحصل الفعلي - (التكلفة + الشحن + العمولة + الضريبة + عمولة التحصيل)
    df['net_profit'] = df['collected_amount'] - (
        df['cost'] + df['shipping'] + df['commission'] + df['tax'] + df['total_collection_fee']
    ) - df['returned_amount']
    # Difference = Original - Net collected (represents collection fees theoretically)
    df['difference'] = df['original_collected'] - df['collected_amount']
    
    # Add Physical Receipt text (نعم / لا)
    df['physical_receipt'] = df['physical_receipt_flag'].apply(lambda x: 'نعم' if x > 0 else 'لا')
    
    # Add Status Column
    def get_status(row):
        if row['return_count'] > 0 and row['transaction_count'] == 0:
            return 'مرتجع'
        if row['collected_amount'] == 0:
            return 'غير مدفوع'
        elif row['collected_amount'] > row['expected_amount'] + 0.1:
            return 'زيادة في التحصيل'
        else:
            return 'مدفوع'

    df['status'] = df.apply(get_status, axis=1)
    
    # Reorder columns for clarity
    display_cols = [
        'order_id', 'platform', 'account_name', 'order_date', 'week_number',
        'expected_amount', 'original_collected', 'total_collection_fee', 'collected_amount',
        'returned_amount', 'difference',
        'cost', 'shipping', 'commission', 'tax', 'net_profit',
        'status', 'physical_receipt', 'physical_receipt_flag', 'items_summary', 'payment_method', 'salla_status', 'last_collection_date', 'transaction_count', 'return_count'
    ]
    df = df[display_cols]

    # Arabic column names for display
    col_names_ar = {
        'order_id':            'رقم الطلب',
        'platform':            'المنصة',
        'account_name':        'الحساب',
        'order_date':          'تاريخ الطلب',
        'week_number':         'رقم الأسبوع',
        'expected_amount':     'المبلغ المتوقع',
        'original_collected':  'المبلغ الأصلي المحصل',
        'total_collection_fee':'عمولة التحصيل',
        'collected_amount':    'صافي المحصل',
        'returned_amount':     'المرتجعات',
        'difference':          'الفرق (عمولات)',
        'cost':                'التكلفة',
        'shipping':            'الشحن',
        'commission':          'العمولة',
        'tax':                 'الضريبة',
        'net_profit':          'صافي الربح',
        'status':              'الحالة',
        'physical_receipt':    'الاستلام الفعلي',
        'items_summary':       'الأصناف',
        'payment_method':      'طريقة الدفع',
        'salla_status':        'حالة سلة',
        'last_collection_date':'آخر تحصيل',
        'transaction_count':   'عدد الحركات',
        'return_count':        'عدد المرتجعات'
    }
    df_display = df.rename(columns=col_names_ar)

    # --- Summary Stats ---
    total_expected   = df['expected_amount'].sum()
    total_original   = df['original_collected'].sum()
    total_fees       = df['total_collection_fee'].sum()
    total_collected  = df['collected_amount'].sum()
    total_returned   = df['returned_amount'].sum()
    total_net_profit = df['net_profit'].sum()
    collection_rate  = (total_collected / total_expected * 100) if total_expected > 0 else 0
    paid_count       = (df['status'] == 'مدفوع').sum()
    unpaid_count     = (df['status'] == 'غير مدفوع').sum()
    returned_count   = (df['status'] == 'مرتجع').sum()
    total_orders     = len(df)

    # Generate Filename
    current_date = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    safe_label   = label.replace(' ', '_').replace('/', '-') if label else f"snapshot_{snapshot_id}"
    report_path  = os.path.join(OUTPUT_DIR, f"Report_{safe_label}_{current_date}.xlsx")
    
    # --- Excel Styles ---
    header_fill    = PatternFill("solid", fgColor="1F3864")   # Dark navy
    subheader_fill = PatternFill("solid", fgColor="2E75B6")   # Blue
    green_fill     = PatternFill("solid", fgColor="E2EFDA")   # Light green
    red_fill       = PatternFill("solid", fgColor="FCE4D6")   # Light red
    yellow_fill    = PatternFill("solid", fgColor="FFF2CC")   # Light yellow
    white_fill     = PatternFill("solid", fgColor="FFFFFF")
    
    header_font    = Font(bold=True, color="FFFFFF", size=12)
    title_font     = Font(bold=True, color="1F3864", size=14)
    kpi_font       = Font(bold=True, color="1F3864", size=22)
    label_font     = Font(bold=True, color="595959", size=10)
    
    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right_align  = Alignment(horizontal='right', vertical='center')

    with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
        
        # =====================================================
        # SHEET 1: Executive Dashboard (الملخص التنفيذي)
        # =====================================================
        # Write a placeholder df to create the sheet
        pd.DataFrame().to_excel(writer, sheet_name='الملخص التنفيذي', index=False)
        ws = writer.sheets['الملخص التنفيذي']
        ws.sheet_view.rightToLeft = True

        # Title
        ws.merge_cells('B2:H2')
        title_cell = ws['B2']
        title_cell.value = f"تقرير المطابقة الأسبوعي - {current_date}"
        title_cell.font = Font(bold=True, color="1F3864", size=18)
        title_cell.alignment = center_align
        title_cell.fill = PatternFill("solid", fgColor="D6E4F0")

        # KPI Cards layout: Row 4 onwards
        kpis = [
            ("إجمالي المبيعات",        f"{total_expected:,.2f} ر.س",   "D6E4F0", "1F3864"),
            ("إجمالي المحصل (صافي)",   f"{total_collected:,.2f} ر.س",  "E2EFDA", "375623"),
            ("عمولات التحصيل",          f"{total_fees:,.2f} ر.س",       "FCE4D6", "C00000"),
            ("إجمالي المرتجعات",        f"{total_returned:,.2f} ر.س",   "FFF2CC", "7F6000"),
            ("صافي الربح الإجمالي",    f"{total_net_profit:,.2f} ر.س", "EBF3FB", "1F3864"),
            ("نسبة التحصيل",           f"{collection_rate:.1f}%",       "E2EFDA", "375623"),
        ]
        
        col_positions = [2, 4, 6, 8, 10, 12]  # B, D, F, H, J, L
        for i, (label, value, bg, fg) in enumerate(kpis):
            col = col_positions[i]
            col_letter = get_column_letter(col)
            next_col   = get_column_letter(col + 1)
            
            # Merge 2 columns for each KPI
            ws.merge_cells(f'{col_letter}4:{next_col}4')
            ws.merge_cells(f'{col_letter}5:{next_col}5')
            ws.merge_cells(f'{col_letter}6:{next_col}6')
            
            label_cell = ws[f'{col_letter}4']
            label_cell.value = label
            label_cell.font = Font(bold=True, color="595959", size=11)
            label_cell.alignment = center_align
            label_cell.fill = PatternFill("solid", fgColor=bg)
            
            value_cell = ws[f'{col_letter}5']
            value_cell.value = value
            value_cell.font = Font(bold=True, color=fg, size=20)
            value_cell.alignment = center_align
            value_cell.fill = PatternFill("solid", fgColor=bg)
            
            ws[f'{col_letter}6'].fill = PatternFill("solid", fgColor=bg)
            
            # Set row heights
            ws.row_dimensions[4].height = 25
            ws.row_dimensions[5].height = 45
            ws.row_dimensions[6].height = 10

        # Order Status Summary (Row 8)
        ws.merge_cells('B8:I8')
        ws['B8'].value = "ملخص حالات الطلبات"
        ws['B8'].font = Font(bold=True, color="FFFFFF", size=12)
        ws['B8'].fill = PatternFill("solid", fgColor="2E75B6")
        ws['B8'].alignment = center_align
        ws.row_dimensions[8].height = 25

        status_data = [
            ("إجمالي الطلبات",    total_orders,    "D6E4F0"),
            ("مدفوع",             paid_count,      "E2EFDA"),
            ("غير مدفوع",         unpaid_count,    "FCE4D6"),
            ("مرتجع",             returned_count,  "F2DCDB"),
        ]
        
        for i, (label, val, bg) in enumerate(status_data):
            col = get_column_letter(2 + i * 2)
            next_col = get_column_letter(3 + i * 2)
            ws.merge_cells(f'{col}9:{next_col}9')
            ws.merge_cells(f'{col}10:{next_col}10')
            ws[f'{col}9'].value = label
            ws[f'{col}9'].font = Font(bold=True, color="595959", size=10)
            ws[f'{col}9'].alignment = center_align
            ws[f'{col}9'].fill = PatternFill("solid", fgColor=bg)
            ws[f'{col}10'].value = val
            ws[f'{col}10'].font = Font(bold=True, color="1F3864", size=18)
            ws[f'{col}10'].alignment = center_align
            ws[f'{col}10'].fill = PatternFill("solid", fgColor=bg)
            ws.row_dimensions[9].height = 22
            ws.row_dimensions[10].height = 40

        # Physical Returns Warning (Row 12)
        # Warning 1: Returned financially but NOT physically received (Risk of loss)
        missing_physical = df[(df['status'] == 'مرتجع') & (df['physical_receipt_flag'] == 0)].shape[0]
        # Warning 2: Physically received but NOT marked as returned financially (Need to process refund in system)
        missed_refund = df[(df['physical_receipt_flag'] > 0) & (df['status'] != 'مرتجع')].shape[0]
        
        ws.merge_cells('B12:I12')
        ws['B12'].value = "تأكيدات المرتجعات (Physical vs Financial)"
        ws['B12'].font = Font(bold=True, color="FFFFFF", size=12)
        ws['B12'].fill = PatternFill("solid", fgColor="C00000") # Red Header
        ws['B12'].alignment = center_align
        ws.row_dimensions[12].height = 25
        
        warnings = [
            ("في الكشوفات وغير مستلم فعلياً", missing_physical, "FCE4D6", "C00000"),
            ("مستلم فعلياً وغير موجود بالكشوفات", missed_refund, "FFF2CC", "7F6000")
        ]
        
        for i, (label, val, bg, fg) in enumerate(warnings):
            col = get_column_letter(2 + i * 4)       # B, F
            next_col = get_column_letter(5 + i * 4)  # E, I
            ws.merge_cells(f'{col}13:{next_col}13')
            ws.merge_cells(f'{col}14:{next_col}14')
            ws[f'{col}13'].value = label
            ws[f'{col}13'].font = Font(bold=True, color="595959", size=10)
            ws[f'{col}13'].alignment = center_align
            ws[f'{col}13'].fill = PatternFill("solid", fgColor=bg)
            
            ws[f'{col}14'].value = val
            ws[f'{col}14'].font = Font(bold=True, color=fg, size=18)
            ws[f'{col}14'].alignment = center_align
            ws[f'{col}14'].fill = PatternFill("solid", fgColor=bg)
            
            ws.row_dimensions[13].height = 22
            ws.row_dimensions[14].height = 40

        # Platform breakdown (Row 16)
        ws.merge_cells('B16:I16')
        ws['B16'].value = "التفصيل حسب المنصة"
        ws['B16'].font = Font(bold=True, color="FFFFFF", size=12)
        ws['B16'].fill = PatternFill("solid", fgColor="2E75B6")
        ws['B16'].alignment = center_align
        ws.row_dimensions[16].height = 25

        platform_summary = df.groupby(['platform', 'account_name']).agg(
            orders=('order_id', 'count'),
            expected=('expected_amount', 'sum'),
            original=('original_collected', 'sum'),
            fees=('total_collection_fee', 'sum'),
            collected=('collected_amount', 'sum'),
            returned=('returned_amount', 'sum'),
            net_profit=('net_profit', 'sum')
        ).reset_index()

        headers_ps = ['المنصة', 'الحساب', 'عدد الطلبات', 'المبلغ المتوقع',
                      'المبلغ الأصلي', 'عمولة التحصيل', 'صافي المحصل', 'المرتجعات', 'صافي الربح']
        for j, h in enumerate(headers_ps):
            cell = ws.cell(row=17, column=2+j)
            cell.value = h
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill("solid", fgColor="1F3864")
            cell.alignment = center_align
            cell.border = thin_border
        ws.row_dimensions[17].height = 22

        for r_idx, row in platform_summary.iterrows():
            row_num = 18 + r_idx
            vals = [row['platform'], row['account_name'], row['orders'],
                    f"{row['expected']:,.2f}", f"{row['original']:,.2f}",
                    f"{row['fees']:,.2f}", f"{row['collected']:,.2f}",
                    f"{row['returned']:,.2f}", f"{row['net_profit']:,.2f}"]
            for j, v in enumerate(vals):
                cell = ws.cell(row=row_num, column=2+j)
                cell.value = v
                cell.alignment = center_align
                cell.border = thin_border
                cell.fill = PatternFill("solid", fgColor="F2F2F2" if r_idx % 2 == 0 else "FFFFFF")
            ws.row_dimensions[row_num].height = 20

        # Column widths
        for col in range(1, 16):
            ws.column_dimensions[get_column_letter(col)].width = 20

        # =====================================================
        # SHEET 2: All Orders (جميع الطلبات)
        # =====================================================
        df_display.to_excel(writer, sheet_name='جميع الطلبات', index=False)
        ws2 = writer.sheets['جميع الطلبات']
        ws2.sheet_view.rightToLeft = True
        _style_data_sheet(ws2, df, header_fill, header_font, thin_border, center_align)

        # =====================================================
        # SHEET 3: Per Platform + Account
        # =====================================================
        for (platform, account), p_df in df.groupby(['platform', 'account_name']):
            sheet_name = f"{platform[:20]}-{account[:8]}" if account else platform[:31]
            sheet_name = sheet_name[:31]  # Excel sheet name limit
            p_display = p_df.rename(columns=col_names_ar)
            p_display.to_excel(writer, sheet_name=sheet_name, index=False)
            ws_p = writer.sheets[sheet_name]
            ws_p.sheet_view.rightToLeft = True
            _style_data_sheet(ws_p, p_df, subheader_fill, header_font, thin_border, center_align)

        # =====================================================
        # SHEET 4: Returns Only (المرتجعات)
        # =====================================================
        returns_df = df[df['returned_amount'] > 0].rename(columns=col_names_ar)
        if not returns_df.empty:
            returns_df.to_excel(writer, sheet_name='المرتجعات', index=False)
            ws_ret = writer.sheets['المرتجعات']
            ws_ret.sheet_view.rightToLeft = True
            _style_data_sheet(ws_ret, df[df['returned_amount'] > 0],
                              PatternFill("solid", fgColor="F2DCDB"), header_font, thin_border, center_align)

        # =====================================================
        # SHEET 5: Items Aggregated Summary (ملخص الأصناف المجمّع)
        # تحليل items_summary وتجميع الكميات لكل صنف لكل حساب
        # =====================================================
        items_rows_df = df[df['items_summary'].str.strip() != ''][
            ['platform', 'account_name', 'items_summary']
        ].copy()

        if not items_rows_df.empty:
            import re as _re
            aggregated = {}  # key: (platform, account, item_name) -> total_qty

            for _, row in items_rows_df.iterrows():
                platform_val = row['platform']
                account_val  = row['account_name'] or 'غير محدد'
                summary      = str(row['items_summary'])

                # تقسيم على ' | ' للطلبات متعددة الأصناف
                parts = summary.split(' | ')
                for part in parts:
                    part = part.strip()
                    if not part:
                        continue
                    # استخراج الاسم والكمية: "اسم المنتج x3"
                    m = _re.match(r'^(.+?)\s+x(\d+)$', part, _re.IGNORECASE)
                    if m:
                        item_name = m.group(1).strip()
                        qty       = int(m.group(2))
                    else:
                        item_name = part
                        qty       = 1

                    key = (platform_val, account_val, item_name)
                    aggregated[key] = aggregated.get(key, 0) + qty

            if aggregated:
                agg_records = [
                    {'المنصة': k[0], 'الحساب': k[1], 'اسم الصنف': k[2], 'إجمالي الكمية': v}
                    for k, v in sorted(aggregated.items(), key=lambda x: (x[0][0], x[0][1], -x[1]))
                ]
                agg_df = pd.DataFrame(agg_records)

                agg_df.to_excel(writer, sheet_name='ملخص الأصناف', index=False)
                ws_agg = writer.sheets['ملخص الأصناف']
                ws_agg.sheet_view.rightToLeft = True

                # تنسيق الورقة
                hdr_fill_items = PatternFill("solid", fgColor="1F3864")
                for cell in ws_agg[1]:
                    cell.font      = Font(bold=True, color="FFFFFF", size=11)
                    cell.fill      = hdr_fill_items
                    cell.alignment = center_align
                    cell.border    = thin_border
                ws_agg.row_dimensions[1].height = 25

                prev_platform = prev_account = None
                for r_idx, row in enumerate(ws_agg.iter_rows(min_row=2), start=2):
                    platform_cell = ws_agg.cell(row=r_idx, column=1).value
                    account_cell  = ws_agg.cell(row=r_idx, column=2).value
                    # تلوين متناوب حسب الحساب
                    bg = "EBF3FB" if (platform_cell, account_cell) != (prev_platform, prev_account) and \
                         (r_idx % 2 == 0) else "FFFFFF"
                    if platform_cell != prev_platform or account_cell != prev_account:
                        bg = "D6E4F0"
                        prev_platform = platform_cell
                        prev_account  = account_cell
                    for cell in row:
                        cell.alignment = center_align
                        cell.border    = thin_border
                        cell.fill      = PatternFill("solid", fgColor=bg)
                    ws_agg.row_dimensions[r_idx].height = 20

                # عرض الأعمدة
                ws_agg.column_dimensions['A'].width = 15
                ws_agg.column_dimensions['B'].width = 15
                ws_agg.column_dimensions['C'].width = 55
                ws_agg.column_dimensions['D'].width = 18

        # =====================================================
        # SHEET 6: Pending & Discrepancies (متأخرات وفروقات)
        # =====================================================
        pending = df[df['status'] != 'مدفوع'].rename(columns=col_names_ar)
        pending.to_excel(writer, sheet_name='متأخرات وفروقات', index=False)
        ws4 = writer.sheets['متأخرات وفروقات']
        ws4.sheet_view.rightToLeft = True
        _style_data_sheet(ws4, df[df['status'] != 'مدفوع'],
                          PatternFill("solid", fgColor="C00000"), header_font, thin_border, center_align)

        # =====================================================
        # SHEET 7: Payment Methods Performance (أداء طرق الدفع)
        # =====================================================
        # Calculate performance per payment method
        pm_df = df.copy()
        
        # Identify cancelled/returned orders for cancellation rate
        # Here we consider returned > 0 or salla_status containing specific keywords as cancelled/returned
        def is_cancelled(row):
            if row['return_count'] > 0:
                return True
            status = str(row.get('salla_status', '')).strip().lower()
            return 'ملغي' in status or 'مسترجع' in status or 'canceled' in status or 'cancelled' in status

        pm_df['is_cancelled'] = pm_df.apply(is_cancelled, axis=1)
        
        # Group by payment method
        # Filter out empty or 'Unknown' payment methods for clearer reporting if needed, 
        # but here we keep all and group them.
        pm_summary = pm_df.groupby('payment_method').agg(
            total_orders=('order_id', 'count'),
            expected_amount=('expected_amount', 'sum'),
            cancelled_orders=('is_cancelled', 'sum'),
            avg_order_value=('expected_amount', 'mean'),
            assigned_commission=('commission', 'sum') # Or you can get it from another place, using sum of commission recorded
        ).reset_index()

        total_orders_all = pm_summary['total_orders'].sum()
        
        # Calculate percentages
        pm_summary['order_percentage'] = (pm_summary['total_orders'] / total_orders_all * 100).fillna(0)
        pm_summary['cancellation_rate'] = (pm_summary['cancelled_orders'] / pm_summary['total_orders'] * 100).fillna(0)
        
        # Sort by total orders descending
        pm_summary = pm_summary.sort_values('total_orders', ascending=False)
        
        # Prepare for Display
        headers_pm = ['طريقة الدفع', 'إجمالي عدد الطلبات', 'نسبة الطلبات (%)', 'إجمالي المبيعات (المتوقع)', 
                      'متوسط قيمة الطلب', 'الطلبات الملغية / المرتجعة', 'معدل الإلغاء (%)', 'عمولة الدفع المعينة']
        
        # Create Sheet
        pd.DataFrame().to_excel(writer, sheet_name='أداء طرق الدفع', index=False)
        ws_pm = writer.sheets['أداء طرق الدفع']
        ws_pm.sheet_view.rightToLeft = True

        # Header Row Formatting
        for j, h in enumerate(headers_pm):
            cell = ws_pm.cell(row=1, column=1+j)
            cell.value = h
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill("solid", fgColor="1F3864")
            cell.alignment = center_align
            cell.border = thin_border
        ws_pm.row_dimensions[1].height = 25

        # Data Rows Formatting
        for r_idx, row in pm_summary.iterrows():
            row_num = 2 + r_idx
            
            pay_method = row['payment_method'] if pd.notna(row['payment_method']) and str(row['payment_method']).strip() else 'غير محدد'
            
            vals = [
                pay_method,
                row['total_orders'],
                f"{row['order_percentage']:.1f}%",
                round(row['expected_amount'], 2),
                round(row['avg_order_value'], 2),
                row['cancelled_orders'],
                f"{row['cancellation_rate']:.1f}%",
                round(row['assigned_commission'], 2)
            ]
            
            for j, v in enumerate(vals):
                cell = ws_pm.cell(row=row_num, column=1+j)
                cell.value = v
                cell.alignment = center_align
                cell.border = thin_border
                cell.fill = PatternFill("solid", fgColor="F2F2F2" if r_idx % 2 == 0 else "FFFFFF")
            ws_pm.row_dimensions[row_num].height = 20

        # Adjust Column Widths
        for col in range(1, len(headers_pm) + 1):
            ws_pm.column_dimensions[get_column_letter(col)].width = 20

    # ── Save KPIs to weekly_reports table ──────────────────
    try:
        conn2 = get_db_connection()
        wk_num = int(df['week_number'].mode()[0]) if not df['week_number'].isna().all() else 0
        yr_num = int(df['order_date'].str[:4].mode()[0]) if not df['order_date'].isna().all() else datetime.now().year
        conn2.execute('''
            INSERT OR REPLACE INTO weekly_reports
            (snapshot_id, week_number, year, label,
             total_orders, total_sales, total_collected, total_uncollected,
             net_profit, collection_rate,
             paid_count, unpaid_count, partial_count, report_path)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            snapshot_id, wk_num, yr_num, label,
            total_orders, round(total_expected, 2), round(total_collected, 2),
            round(total_expected - total_collected, 2),
            round(total_net_profit, 2), round(collection_rate, 1),
            int(paid_count), int(unpaid_count), 0,
            report_path
        ))
        conn2.commit()
        conn2.close()
        print(f"  -> KPIs saved to weekly_reports (snapshot_id={snapshot_id})")
    except Exception as e:
        print(f"  -> Warning: could not save to weekly_reports: {e}")

    print(f"✅ Report generated: {report_path}")
    print(f"\n{'='*60}")
    print(f"  إجمالي الطلبات:             {total_orders}")
    print(f"  إجمالي المبيعات:            {total_expected:,.2f} ر.س")
    print(f"  المبلغ الأصلي المحصل:       {total_original:,.2f} ر.س")
    print(f"  عمولات التحصيل (الفرق):     {total_fees:,.2f} ر.س")
    print(f"  صافي المحصل:                {total_collected:,.2f} ر.س")
    print(f"  إجمالي المرتجعات:           {total_returned:,.2f} ر.س")
    print(f"  صافي الربح:                 {total_net_profit:,.2f} ر.س")
    print(f"  نسبة التحصيل:               {collection_rate:.1f}%")
    print(f"{'='*60}")
    print(f"\nتفصيل حسب الحالة:")
    print(df.groupby(['platform', 'account_name', 'status']).size().unstack(fill_value=0))


def _style_data_sheet(ws, source_df, header_fill, header_font, thin_border, center_align):
    """Apply consistent styling to a data worksheet."""
    status_colors = {
        'مدفوع':              'E2EFDA',
        'غير مدفوع':          'FCE4D6',
        'زيادة في التحصيل':   'D6E4F0',
    }
    
    # Style header row
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = 'A2'

    # Style data rows based on status
    status_col_idx = None
    for i, cell in enumerate(ws[1]):
        if cell.value in ('الحالة', 'status'):
            status_col_idx = i + 1
            break

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        status_val = ''
        if status_col_idx:
            status_val = ws.cell(row=row_idx, column=status_col_idx).value or ''
        
        bg = status_colors.get(status_val, 'FFFFFF')
        fill = PatternFill("solid", fgColor=bg)
        
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_align
            if status_val:
                cell.fill = fill
        ws.row_dimensions[row_idx].height = 18

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 35)


if __name__ == "__main__":
    generate_weekly_report()
